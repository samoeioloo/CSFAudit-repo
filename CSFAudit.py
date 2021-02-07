#!/usr/bin/env python3
# -*- coding: UTF-8 -*-
#
# generated by wxGlade 1.0.1 on Sun Feb  7 12:08:36 2021
#

import wx

# begin wxGlade: dependencies
# end wxGlade

# begin wxGlade: extracode
import openpyxl
# end wxGlade


class csFrame(wx.Frame):
    def __init__(self, *args, **kwds):
        # begin wxGlade: csFrame.__init__
        kwds["style"] = kwds.get("style", 0) | wx.DEFAULT_FRAME_STYLE
        wx.Frame.__init__(self, *args, **kwds)
        self.SetSize((800, 450))
        self.SetTitle("Cybersecurity Framework Auditor")

        sizer_1 = wx.BoxSizer(wx.VERTICAL)

        self.nb = wx.Notebook(self, wx.ID_ANY)
        sizer_1.Add(self.nb, 1, wx.EXPAND, 0)

        self.nb1 = wx.Panel(self.nb, wx.ID_ANY)
        self.nb.AddPage(self.nb1, "Main")

        sizer_2 = wx.BoxSizer(wx.VERTICAL)

        label_1 = wx.StaticText(self.nb1, wx.ID_ANY, "Systems under audit")
        sizer_2.Add(label_1, 0, wx.ALL, 4)

        sizer_3 = wx.BoxSizer(wx.HORIZONTAL)
        sizer_2.Add(sizer_3, 0, wx.EXPAND, 0)

        self.lbSystem = wx.ListBox(self.nb1, wx.ID_ANY, choices=["choice 1"], style=wx.LB_HSCROLL | wx.LB_SINGLE | wx.LB_SORT)
        self.lbSystem.SetMinSize((200, 300))
        #Open workbook, add list of sheet names into listbox widget
        self.wb=openpyxl.load_workbook("C:/Users/P/OneDrive/Desktop/Dev Projects/Cybersecurity Framework/CSF-Audit.xlsx")
        self.lbSystem.Clear()
        self.lbSystem.Append(self.wb.sheetnames)
        self.lbSystem.SetSelection(0)
        sizer_3.Add(self.lbSystem, 0, wx.ALL, 4)

        sizer_3.Add((20, 20), 0, 0, 0)

        sizer_4 = wx.BoxSizer(wx.VERTICAL)
        sizer_3.Add(sizer_4, 0, wx.EXPAND, 0)

        self.tcSysInfo = wx.TextCtrl(self.nb1, wx.ID_ANY, "", style=wx.TE_MULTILINE | wx.TE_WORDWRAP)
        self.tcSysInfo.SetMinSize((400, 200))
        sizer_4.Add(self.tcSysInfo, 0, 0, 0)

        sizer_5 = wx.BoxSizer(wx.HORIZONTAL)
        sizer_4.Add(sizer_5, 0, wx.EXPAND, 0)

        self.bAdd = wx.Button(self.nb1, wx.ID_ANY, "Add")
        self.bAdd.SetMinSize((60, 20))
        sizer_5.Add(self.bAdd, 0, wx.ALL, 4)

        self.bDel = wx.Button(self.nb1, wx.ID_ANY, "Delete")
        self.bDel.SetMinSize((60, 20))
        sizer_5.Add(self.bDel, 0, wx.ALL, 4)

        self.bOpen = wx.Button(self.nb1, wx.ID_ANY, "Open")
        self.bOpen.SetMinSize((60, 20))
        sizer_5.Add(self.bOpen, 0, wx.ALL, 4)

        self.bStore = wx.Button(self.nb1, wx.ID_ANY, "Store")
        self.bStore.SetMinSize((60, 20))
        sizer_5.Add(self.bStore, 0, wx.ALL, 4)

        sizer_6 = wx.BoxSizer(wx.HORIZONTAL)
        sizer_4.Add(sizer_6, 0, wx.EXPAND, 0)

        label_3 = wx.StaticText(self.nb1, wx.ID_ANY, "System Name")
        label_3.SetMinSize((60, 35))
        sizer_6.Add(label_3, 0, wx.ALL, 4)

        self.tcSystem = wx.TextCtrl(self.nb1, wx.ID_ANY, "")
        self.tcSystem.SetMinSize((120, 20))
        sizer_6.Add(self.tcSystem, 0, wx.ALL, 4)

        static_line_1 = wx.StaticLine(self.nb1, wx.ID_ANY)
        sizer_2.Add(static_line_1, 0, wx.ALL | wx.EXPAND, 4)

        label_2 = wx.StaticText(self.nb1, wx.ID_ANY, "Select a system and press Open or double click, go to the Audit page and work on it. Press \"Delete\" to remove the currently selected system. Press \"Add\" to create a new system to audit. To save changes back to disk, select \"Store\".", style=wx.ALIGN_LEFT)
        label_2.SetMinSize((600, 32))
        sizer_2.Add(label_2, 0, wx.ALL, 4)

        self.nb2 = wx.Panel(self.nb, wx.ID_ANY)
        self.nb.AddPage(self.nb2, "Audit")

        sizer_7 = wx.BoxSizer(wx.VERTICAL)

        self.tree = wx.TreeCtrl(self.nb2, wx.ID_ANY)
        self.tree.SetMinSize((600, 200))
        self.root = self.tree.AddRoot("CSF")
        sizer_7.Add(self.tree, 0, wx.ALL | wx.EXPAND, 1)

        sizer_8 = wx.BoxSizer(wx.HORIZONTAL)
        sizer_7.Add(sizer_8, 0, wx.EXPAND, 0)

        sizer_9 = wx.BoxSizer(wx.VERTICAL)
        sizer_8.Add(sizer_9, 1, wx.EXPAND, 0)

        label_4 = wx.StaticText(self.nb2, wx.ID_ANY, "Evidence")
        sizer_9.Add(label_4, 0, wx.LEFT | wx.RIGHT | wx.TOP, 4)

        self.tcEvidence = wx.TextCtrl(self.nb2, wx.ID_ANY, "", style=wx.TE_MULTILINE | wx.TE_WORDWRAP)
        self.tcEvidence.SetMinSize((300, 150))
        sizer_9.Add(self.tcEvidence, 0, wx.ALL, 4)

        sizer_11 = wx.BoxSizer(wx.VERTICAL)
        sizer_8.Add(sizer_11, 0, wx.EXPAND, 0)

        self.tcShort = wx.TextCtrl(self.nb2, wx.ID_ANY, "", style=wx.TE_READONLY)
        sizer_11.Add(self.tcShort, 0, wx.ALIGN_CENTER_HORIZONTAL | wx.ALL, 4)

        self.chRating = wx.Choice(self.nb2, wx.ID_ANY, choices=["High", "Medium", "Low", "Compliant", "Unknown"])
        self.chRating.SetSelection(4)
        sizer_11.Add(self.chRating, 0, wx.ALIGN_CENTER_HORIZONTAL | wx.ALL, 4)

        sizer_11.Add((20, 70), 0, wx.ALIGN_CENTER_HORIZONTAL, 0)

        self.bSave = wx.Button(self.nb2, wx.ID_ANY, "Save")
        self.bSave.SetMinSize((60, 20))
        sizer_11.Add(self.bSave, 0, wx.ALIGN_CENTER_HORIZONTAL | wx.ALL, 4)

        sizer_10 = wx.BoxSizer(wx.VERTICAL)
        sizer_8.Add(sizer_10, 1, wx.EXPAND, 0)

        label_5 = wx.StaticText(self.nb2, wx.ID_ANY, "Assessment")
        sizer_10.Add(label_5, 0, wx.LEFT | wx.RIGHT | wx.TOP, 4)

        self.tcAssessment = wx.TextCtrl(self.nb2, wx.ID_ANY, "", style=wx.TE_MULTILINE | wx.TE_WORDWRAP)
        self.tcAssessment.SetMinSize((300, 150))
        sizer_10.Add(self.tcAssessment, 0, wx.ALL, 4)

        self.nb2.SetSizer(sizer_7)

        self.nb1.SetSizer(sizer_2)

        self.SetSizer(sizer_1)

        self.Layout()

        self.Bind(wx.EVT_LISTBOX, self.onListChange, self.lbSystem)
        self.Bind(wx.EVT_LISTBOX_DCLICK, self.onOpenPressed, self.lbSystem)
        self.Bind(wx.EVT_BUTTON, self.onAddPressed, self.bAdd)
        self.Bind(wx.EVT_BUTTON, self.onDelPressed, self.bDel)
        self.Bind(wx.EVT_BUTTON, self.onOpenPressed, self.bOpen)
        self.Bind(wx.EVT_BUTTON, self.onStorePressed, self.bStore)
        self.Bind(wx.EVT_TREE_SEL_CHANGED, self.onTreeSelected, self.tree)
        self.Bind(wx.EVT_BUTTON, self.onSavePressed, self.bSave)
        # end wxGlade
        self.displayInfo()

    def wsRead(self, name):
        #tree control is populated for a system
        self.tree.DeleteChildren(self.root) #delete everything from root so that there is a clean tree to populate
        
        #set group and category to blank bc as we build tree, start a new branch to every change of group and new sub-branch for every change of category
        group = "" 
        category = ""

        self.ws = self.wb[name]
        i=4 #start row for reading controls
        while True:
            cell1 = "B" + str(i)
            cell2 = "D" + str(i)
            cell3 = "G" + str(i)
            if self.ws[cell1].value == "END": # tree is complete
                break
            if self.ws[cell1].value != group: #group in the correct record has changed?
                #append new group branch to tree and made current group
                grpId = self.tree.AppendItem(self.root, self.ws[cell1].value) 
                group = self.ws[cell1].value
                #force a new category subbranch to be sent
                category = ""
            if self.ws[cell2].value != category:
                #append new category
                catId = self.tree.AppendItem(grpId, self.ws[cell2].value)
                category = self.ws[cell2].value
            #add leaf to tree
            self.tree.AppendItem(catId, self.ws[cell3].value, data=(i)) #data here is the spreadsheet line number used for control
            i += 1
        idx = self.lbSystem.FindString(name) #obtain list box index
        self.lbSystem.SetSelection(idx)
        self.nb.SetPageText(1, name)
        return

    def wsRemove(self, name):
        self.ws = self.wb[name]
        self.wb.remove(self.ws)
        self.wb.save("C:/Users/P/OneDrive/Desktop/Dev Projects/Cybersecurity Framework/CSF-Audit.xlsx")
        return

    def wsWrite(self):
          # saves contents of systems summary to audit sheet E2 cell of currently selected worksheet
            self.ws['E2'] = self.tcSysInfo.GetValue()
            self.wb.save("C:/Users/P/OneDrive/Desktop/Dev Projects/Cybersecurity Framework/CSF-Audit.xlsx")
            return

    #get and display info
    def displayInfo(self):
        #get system name
        listNo = self.lbSystem.GetSelection()
        listName = self.lbSystem.GetString(listNo)
        # use list name to set sheet to active
        self.ws = self.wb[listName]
        self.tcSysInfo.SetValue(self.ws['E2'].value)
        return

    def onListChange(self, event):  # wxGlade: csFrame.<event_handler>
        self.displayInfo()
        event.Skip()


    def onAddPressed(self, event):  # wxGlade: csFrame.<event_handler>
        # check system name
        sysname = self.tcSystem.GetValue()
        if len(sysname) == 0: # if blank
            wx.MessageBox("Please enter a system name", "Error", wx.OK)
        elif self.lbSystem.FindString(sysname) != wx.NOT_FOUND:  # if system name is already in list
            wx.MessageBox("System already under audit", "Error", wx.OK)
        # append system to list
        else:
            self.lbSystem.Append(sysname)
            self.tcSystem.SetValue("")
            # create ws in wb (do this by copying)
            source = self.wb['Template']
            self.ws = self.wb.copy_worksheet(source)
            # change title of sheet to new system name
            self.ws.title = sysname
            #write back as blank audit record into wb
            self.wb.save("C:/Users/P/OneDrive/Desktop/Dev Projects/Cybersecurity FrameworkCSF-Audit.xlsx")

            # read new ws into application, ready for audit
            self.wsRead(sysname)

        event.Skip()
    def onOpenPressed(self, event):  # wxGlade: csFrame.<event_handler>
        listNo = self.lbSystem.GetSelection()
        listName = self.lbSystem.GetString(listNo)
        self.wsRead(listName)
        event.Skip()
        
    def onDelPressed(self, event):  # wxGlade: csFrame.<event_handler>
        listNo = self.lbSystem.GetSelection()
        listName = self.lbSystem.GetString(listNo)
        if listName == "Template":
            wx.MessageBox("You cannot delete the Template", "Error", wx.OK)
        else:
            dlg = wx.MessageDialog(None, "Do you want to delete " + listName + "? ", \
                "Confirm",wx.YES_NO)
            if dlg.ShowModal() == wx.ID_YES:
                self.wsRemove(listName)
                # if system is the currently loaded system
            if self.nb.GetPageText(1) == listName: 
                listNo = self.lbSystem.FindString("Template") #makes Template the currently selected item
                self.lbSystem.SetSelection(listNo)
                self.wsRead("Template") #load 2nd tab
        event.Skip()

    def onStorePressed(self, event):  # wxGlade: csFrame.<event_handler>
        #this function sends the changes back to disk
        self.wsWrite()
        event.Skip()
    

    def onTreeSelected(self, event):  # wxGlade: csFrame.<event_handler>
        reqNo = self.tree.GetSelection() # get index of selected requirement
        if not self.tree.ItemHasChildren(reqNo): # check if selected item is a leaf
            regrow = self.tree.GetItemData(reqNo)
            evidence = "H" + str(regrow)
            compliance = "I" + str(regrow)
            assessment = "J" +str(regrow)
            if self.ws[evidence].value == None:
              self.tcEvidence.SetValue("")
            else:
                self.tcEvidence.SetValue(self.ws[evidence].value)
            if self.ws[compliance].value == None:
                self.chRating.SetSelection(4)
            else:
                choice = self.chRating.FindString(self.ws[compliance].value) # identify index of current compliance rating value
                self.chRating.SetSelection(choice) # set value in choice widget
            if self.ws[assessment].value == None:
                self.tcAssessment.SetValue("")
            else:
                self.tcAssessment.SetValue(self.ws[assessment].value)
            #contatenate cols A,C,F to form short identifier for the requirement. display in entry box
            short = self.ws["A"+str(regrow)].value+"."+ self.ws["C"+str(regrow)].value+"."+str(self.ws["F"+str(regrow)].value)
            self.tcShort.SetValue(short)
        event.Skip()

    def onSavePressed(self, event):  # wxGlade: csFrame.<event_handler>
        # get spreadsheet row of current requirement
        reqno = self.tree.GetSelection()
        regrow = self.tree.GetItemData(reqno)

        evidence = "H"+str(regrow)
        compliance = "I"+str(regrow)
        assessment = "J"+str(regrow)

        # write evidence data into spreadsheet cells
        self.ws[evidence] = self.tcEvidence.GetValue()

        choice = self.chRating.GetCurrentSelection()
        self.ws[compliance] = self.chRating.GetString(choice)
        self.ws[assessment] = self.tcAssessment.GetValue()
        event.Skip()
        # NB: changes are not stored on disk after pressing save (only saved into workbook loaded into application)
# end of class csFrame

class csfApp(wx.App):
    def OnInit(self):
        self.frame = csFrame(None, wx.ID_ANY, "")
        self.SetTopWindow(self.frame)
        self.frame.Show()
        return True

# end of class csfApp

if __name__ == "__main__":
    app = csfApp(0)
    app.MainLoop()
